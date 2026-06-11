"""Import RSA CMS launch data from CSV/Excel templates into DynamoDB.

Batch 26 is safe by default:
- Running without --execute is a dry run only.
- Existing DynamoDB records are skipped by default.
- --overwrite must be explicitly provided to replace existing records.
- CSV import works with Python standard library only.
- Excel import is optional and requires openpyxl in the local venv.

Examples:
    python scripts/import_launch_data_to_dynamodb.py --all
    python scripts/import_launch_data_to_dynamodb.py --table products
    python scripts/import_launch_data_to_dynamodb.py --all --execute
    python scripts/import_launch_data_to_dynamodb.py --source excel --all
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime, timezone
from decimal import Decimal
import json
from pathlib import Path
import re
import sys
from typing import Any

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

try:
    from dotenv import load_dotenv  # type: ignore
    load_dotenv(BACKEND_ROOT / '.env')
except Exception:
    pass

from app.database import (  # noqa: E402
    get_aws_region,
    get_dynamodb_table,
    get_table_name,
)

DEFAULT_TEMPLATE_DIR = BACKEND_ROOT / 'app' / 'data' / 'import_templates'
DEFAULT_CSV_DIR = DEFAULT_TEMPLATE_DIR / 'csv'
DEFAULT_EXCEL_FILE = DEFAULT_TEMPLATE_DIR / 'rsa_launch_data_template.xlsx'
MANIFEST_PATH = DEFAULT_TEMPLATE_DIR / 'schema_manifest.json'

VALID_FLAGS = {'Y', 'N'}
VALID_CONTACT_TYPES = {'Company Contact', 'Contact Person', 'Social Media'}
DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')
ID_RE = re.compile(r'^([A-Z]{4})-(\d{7})$')

PRIMARY_KEYS = {
    'categories': 'category_id',
    'brands': 'brand_id',
    'key_features': 'key_feat_id',
    'products': 'product_id',
    'services': 'service_id',
    'about': 'about_id',
    'project_gallery': 'project_id',
    'contact_us': 'contact_us_id',
}

DEFAULT_PREFIXES = {
    'categories': 'CATG',
    'brands': 'BRND',
    'key_features': 'KFEA',
    'services': 'SERV',
    'about': 'ABOU',
    'project_gallery': 'PROJ',
}

NUMERIC_FIELDS = {
    'display_seq',
    'price',
    'sale_price',
    'stock_quantity',
    'low_stock_threshold',
    'last_number',
}

INTEGER_FIELDS = {'display_seq', 'stock_quantity', 'low_stock_threshold', 'last_number'}

TABLE_ORDER = [
    'categories',
    'brands',
    'key_features',
    'products',
    'services',
    'about',
    'project_gallery',
    'contact_us',
]


class ImportErrorWithDetails(Exception):
    pass


def load_manifest(path: Path = MANIFEST_PATH) -> dict[str, Any]:
    if not path.exists():
        raise ImportErrorWithDetails(f'Missing schema manifest: {path}')
    return json.loads(path.read_text(encoding='utf-8'))


def clean_string(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


def normalize_key(value: str) -> str:
    return re.sub(r'[^a-z0-9]+', '-', value.strip().lower()).strip('-')


def slugify(value: str) -> str:
    return normalize_key(value)


def normalize_phone_number(value: str) -> str:
    digits = re.sub(r'\D+', '', value or '')
    if digits.startswith('0063'):
        digits = digits[2:]
    if digits.startswith('0639') and len(digits) == 13:
        return '63' + digits[2:]
    if digits.startswith('09') and len(digits) == 11:
        return '63' + digits[1:]
    if digits.startswith('9') and len(digits) == 10:
        return '63' + digits
    return digits


def resolve_media_path(folder: str, filename: str, existing_path: str = '') -> str:
    if existing_path:
        return existing_path
    if not filename:
        return ''
    safe_name = Path(filename).name.replace(' ', '-')
    return f'uploads/{folder}/{safe_name}'


def read_csv_rows(path: Path, expected_headers: list[str]) -> list[dict[str, str]]:
    if not path.exists():
        raise ImportErrorWithDetails(f'Missing CSV file: {path}')
    with path.open('r', newline='', encoding='utf-8-sig') as handle:
        reader = csv.DictReader(handle)
        headers = list(reader.fieldnames or [])
        if headers != expected_headers:
            raise ImportErrorWithDetails(
                f'Headers do not match template for {path.name}.\n'
                f'Expected: {expected_headers}\nActual:   {headers}'
            )
        rows = []
        for row in reader:
            cleaned = {header: clean_string(row.get(header)) for header in expected_headers}
            if any(cleaned.values()):
                rows.append(cleaned)
        return rows


def read_excel_rows(path: Path, sheet_name: str, expected_headers: list[str]) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise ImportErrorWithDetails(
            'Excel import requires openpyxl. Install it in the backend venv with: pip install openpyxl\n'
            'CSV import does not require openpyxl and is available now.'
        ) from exc

    if not path.exists():
        raise ImportErrorWithDetails(f'Missing Excel file: {path}')

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ImportErrorWithDetails(f'Missing Excel sheet {sheet_name!r} in {path}')

    sheet = workbook[sheet_name]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [clean_string(value) for value in next(rows_iter)]
    except StopIteration:
        raise ImportErrorWithDetails(f'Excel sheet {sheet_name!r} is empty.')

    if headers != expected_headers:
        raise ImportErrorWithDetails(
            f'Headers do not match template for sheet {sheet_name}.\n'
            f'Expected: {expected_headers}\nActual:   {headers}'
        )

    rows: list[dict[str, str]] = []
    for raw_row in rows_iter:
        cleaned = {
            header: clean_string(raw_row[index] if index < len(raw_row) else '')
            for index, header in enumerate(expected_headers)
        }
        if any(cleaned.values()):
            rows.append(cleaned)
    return rows


def select_tables(manifest: dict[str, Any], selected: str) -> list[str]:
    tables = manifest['tables']
    if selected == 'all':
        return [table for table in TABLE_ORDER if table in tables]
    if selected not in tables:
        raise SystemExit(f"Unknown table '{selected}'. Valid tables: all, {', '.join(sorted(tables))}")
    return [selected]


def load_input_rows(args: argparse.Namespace, manifest: dict[str, Any], selected_tables: list[str]) -> dict[str, list[dict[str, str]]]:
    tables = manifest['tables']
    loaded: dict[str, list[dict[str, str]]] = {}

    for table in selected_tables:
        schema = tables[table]
        headers = schema['headers']
        if args.source == 'csv':
            csv_path = Path(args.csv_dir) / schema['file']
            loaded[table] = read_csv_rows(csv_path, headers)
        elif args.source == 'excel':
            loaded[table] = read_excel_rows(Path(args.excel_file), schema['sheet'], headers)
        else:
            raise ImportErrorWithDetails(f'Unsupported source: {args.source}')

    return loaded


def validate_rows(manifest: dict[str, Any], rows_by_table: dict[str, list[dict[str, str]]]) -> list[str]:
    errors: list[str] = []

    for table, rows in rows_by_table.items():
        schema = manifest['tables'][table]
        required = schema.get('required', [])
        for index, row in enumerate(rows, start=2):
            for field in required:
                if not row.get(field):
                    errors.append(f"{table}: row {index} missing required field '{field}'.")

            for flag_field in ('show_flag', 'show_pack_flag'):
                value = row.get(flag_field)
                if value and value not in VALID_FLAGS:
                    errors.append(f"{table}: row {index} field '{flag_field}' must be Y or N.")

            contact_type = row.get('contact_type')
            if contact_type and contact_type not in VALID_CONTACT_TYPES:
                errors.append(f"{table}: row {index} contact_type must be one of: {', '.join(sorted(VALID_CONTACT_TYPES))}.")

            for numeric_field in NUMERIC_FIELDS:
                value = row.get(numeric_field)
                if not value:
                    continue
                try:
                    Decimal(value.replace(',', ''))
                except Exception:
                    errors.append(f"{table}: row {index} field '{numeric_field}' must be numeric.")

            completed_date = row.get('completed_date')
            if completed_date and not DATE_RE.match(completed_date):
                errors.append(f"{table}: row {index} completed_date must use YYYY-MM-DD.")

            phone = row.get('phone_number')
            if phone:
                normalized = normalize_phone_number(phone)
                if not (normalized.startswith('639') and len(normalized) == 12):
                    errors.append(
                        f"{table}: row {index} phone_number should be a Philippines mobile number "
                        f"such as +63 919 123 4567 or 0919 123 4567."
                    )

    return errors


def get_existing_counters(region: str, execute: bool) -> dict[str, int]:
    if not execute:
        return {}

    counters: dict[str, int] = {}
    try:
        table = get_dynamodb_table(get_table_name('id_counters'), region_name=region)
        response = table.scan()
        for item in response.get('Items', []):
            prefix = item.get('id_prefix')
            last_number = item.get('last_number')
            if prefix:
                try:
                    counters[str(prefix)] = int(last_number)
                except Exception:
                    pass
    except Exception:
        # If counters table is unavailable, continue with input-derived counters.
        pass
    return counters


def update_counter_from_id(counters: dict[str, int], record_id: str) -> None:
    match = ID_RE.match(record_id or '')
    if not match:
        return
    prefix = match.group(1)
    number = int(match.group(2))
    counters[prefix] = max(counters.get(prefix, 0), number)


def next_id(counters: dict[str, int], prefix: str) -> str:
    prefix = (prefix or 'MISC').upper()[:4].ljust(4, 'X')
    number = counters.get(prefix, 0) + 1
    counters[prefix] = number
    return f'{prefix}-{number:07d}'


def decimal_or_none(value: str) -> Decimal | None:
    value = clean_string(value).replace(',', '')
    if not value:
        return None
    return Decimal(value)


def number_value(field: str, value: str) -> Any:
    parsed = decimal_or_none(value)
    if parsed is None:
        return None
    if field in INTEGER_FIELDS:
        return int(parsed)
    return parsed


def base_record(row: dict[str, str], fields: list[str]) -> dict[str, Any]:
    item: dict[str, Any] = {}
    for field in fields:
        raw = row.get(field, '')
        if raw == '':
            continue
        if field in NUMERIC_FIELDS:
            item[field] = number_value(field, raw)
        else:
            item[field] = raw
    return item


def build_reference_maps(rows_by_table: dict[str, list[dict[str, str]]], region: str, execute: bool) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    categories: dict[str, dict[str, Any]] = {}
    brands: dict[str, dict[str, Any]] = {}

    for row in rows_by_table.get('categories', []):
        key = row.get('category_key')
        if key:
            categories[key] = {
                'category_id': row.get('category_id'),
                'category_key': key,
                'category_name': row.get('category_name'),
                'category_prefix': row.get('category_prefix'),
            }

    for row in rows_by_table.get('brands', []):
        key = row.get('product_brand_key')
        if key:
            brands[key] = {
                'brand_id': row.get('brand_id'),
                'product_brand_key': key,
                'product_brand_name': row.get('product_brand_name'),
                'brand_logo_path': row.get('brand_logo_path') or resolve_media_path('brands', row.get('brand_logo_filename', '')),
            }

    if execute:
        # Merge existing DynamoDB reference data for product enrichment when importing products only.
        try:
            category_table = get_dynamodb_table(get_table_name('categories'), region_name=region)
            for item in category_table.scan().get('Items', []):
                key = item.get('category_key')
                if key and key not in categories:
                    categories[str(key)] = dict(item)
        except Exception:
            pass

        try:
            brand_table = get_dynamodb_table(get_table_name('brands'), region_name=region)
            for item in brand_table.scan().get('Items', []):
                key = item.get('product_brand_key')
                if key and key not in brands:
                    brands[str(key)] = dict(item)
        except Exception:
            pass

    return categories, brands


def prefix_for_product(row: dict[str, str], categories: dict[str, dict[str, Any]]) -> str:
    category_key = row.get('category_key', '')
    category = categories.get(category_key, {})
    prefix = clean_string(category.get('category_prefix')) or clean_string(row.get('category_prefix'))
    if prefix:
        return prefix.upper()[:4]
    return re.sub(r'[^A-Z0-9]', '', category_key.upper())[:4] or 'PROD'


def transform_rows(rows_by_table: dict[str, list[dict[str, str]]], region: str, execute: bool) -> tuple[dict[str, list[dict[str, Any]]], dict[str, int]]:
    now = datetime.now(timezone.utc).isoformat()
    counters = get_existing_counters(region, execute=execute)

    # Prime counters from provided IDs before generating missing IDs.
    for table, rows in rows_by_table.items():
        pk = PRIMARY_KEYS.get(table)
        if not pk:
            continue
        for row in rows:
            update_counter_from_id(counters, row.get(pk, ''))

    categories, brands = build_reference_maps(rows_by_table, region=region, execute=execute)
    output: dict[str, list[dict[str, Any]]] = {}

    for table, rows in rows_by_table.items():
        items: list[dict[str, Any]] = []
        for row in rows:
            item = base_record(row, list(row.keys()))
            pk = PRIMARY_KEYS[table]

            if not item.get(pk):
                if table == 'products':
                    item[pk] = next_id(counters, prefix_for_product(row, categories))
                elif table == 'contact_us':
                    contact_type = row.get('contact_type')
                    if contact_type == 'Company Contact':
                        item[pk] = 'CONT-0000001'
                        update_counter_from_id(counters, item[pk])
                    elif contact_type == 'Contact Person':
                        item[pk] = next_id(counters, 'CPER')
                    elif contact_type == 'Social Media':
                        item[pk] = next_id(counters, 'SOCM')
                    else:
                        item[pk] = next_id(counters, 'CONT')
                else:
                    item[pk] = next_id(counters, DEFAULT_PREFIXES[table])
            else:
                update_counter_from_id(counters, item[pk])

            # Defaults and derived values.
            item.setdefault('show_flag', 'Y')
            item.setdefault('display_seq', 999)

            if table == 'products':
                item.setdefault('show_pack_flag', 'N')
                category_key = row.get('category_key', '')
                category = categories.get(category_key, {})
                item['category_id'] = item.get('category_id') or category.get('category_id')
                item['category_name'] = item.get('category_name') or category.get('category_name')
                item['category_prefix'] = item.get('category_prefix') or category.get('category_prefix')

                brand_key = row.get('product_brand_key', '')
                brand = brands.get(brand_key, {})
                item['brand_id'] = item.get('brand_id') or brand.get('brand_id')
                item['product_brand_name'] = item.get('product_brand_name') or brand.get('product_brand_name')
                item['brand_logo_path'] = item.get('brand_logo_path') or brand.get('brand_logo_path')

                if not item.get('product_name'):
                    parts = [item.get('product_brand_name'), item.get('feature_01'), item.get('subcategory')]
                    item['product_name'] = ' '.join(str(part) for part in parts if part)

                if not item.get('product_slug') and item.get('product_name'):
                    item['product_slug'] = slugify(str(item['product_name']))
                if not item.get('image_path'):
                    item['image_path'] = resolve_media_path('products', row.get('image_filename', ''))

            elif table == 'brands':
                if not item.get('brand_logo_path'):
                    item['brand_logo_path'] = resolve_media_path('brands', row.get('brand_logo_filename', ''))

            elif table == 'services':
                if not item.get('image_path'):
                    item['image_path'] = resolve_media_path('services', row.get('image_filename', ''))

            elif table == 'about':
                if not item.get('image_path'):
                    item['image_path'] = resolve_media_path('about', row.get('image_filename', ''))

            elif table == 'project_gallery':
                if not item.get('project_slug') and item.get('project_title'):
                    item['project_slug'] = slugify(str(item['project_title']))
                if not item.get('image_path'):
                    item['image_path'] = resolve_media_path('project-gallery', row.get('image_filename', ''))

            elif table == 'contact_us':
                if item.get('phone_number'):
                    item['phone_number_normalized'] = normalize_phone_number(str(item['phone_number']))
                if item.get('contact_type') == 'Contact Person' and not item.get('photo_path'):
                    item['photo_path'] = resolve_media_path('contact-persons', row.get('photo_filename', ''))
                if item.get('contact_type') != 'Contact Person':
                    item.pop('photo_filename', None)
                    item.pop('photo_path', None)

            item.setdefault('created_at', now)
            item['updated_at'] = now
            item.setdefault('created_by', 'launch-import')
            item['updated_by'] = 'launch-import'

            items.append(remove_empty_values(item))

        output[table] = items

    return output, counters


def remove_empty_values(value: Any) -> Any:
    if isinstance(value, dict):
        cleaned: dict[str, Any] = {}
        for key, inner in value.items():
            if inner is None or inner == '':
                continue
            cleaned_inner = remove_empty_values(inner)
            if cleaned_inner == {} or cleaned_inner == []:
                continue
            cleaned[key] = cleaned_inner
        return cleaned
    if isinstance(value, list):
        return [remove_empty_values(item) for item in value if item not in (None, '')]
    return value


def primary_key_for_table(table: str) -> str:
    return PRIMARY_KEYS[table]


def get_existing_item(table_name: str, pk_field: str, pk_value: str, region: str) -> dict[str, Any] | None:
    table = get_dynamodb_table(table_name, region_name=region)
    return table.get_item(Key={pk_field: pk_value}).get('Item')


def put_items(transformed: dict[str, list[dict[str, Any]]], counters: dict[str, int], region: str, overwrite: bool, sync_counters: bool) -> dict[str, dict[str, int]]:
    results: dict[str, dict[str, int]] = {}

    for table_name_logical, items in transformed.items():
        table_name = get_table_name(table_name_logical)
        table = get_dynamodb_table(table_name, region_name=region)
        pk_field = primary_key_for_table(table_name_logical)
        written = 0
        skipped = 0

        for item in items:
            existing = get_existing_item(table_name, pk_field, item[pk_field], region)
            if existing and not overwrite:
                skipped += 1
                continue
            table.put_item(Item=item)
            written += 1

        results[table_name_logical] = {
            'table_name': table_name,
            'written': written,
            'skipped_existing': skipped,
            'total_input': len(items),
        }

    if sync_counters:
        counter_table = get_dynamodb_table(get_table_name('id_counters'), region_name=region)
        now = datetime.now(timezone.utc).isoformat()
        counter_writes = 0
        for prefix, last_number in sorted(counters.items()):
            counter_table.put_item(Item={
                'id_prefix': prefix,
                'last_number': int(last_number),
                'updated_at': now,
                'updated_by': 'launch-import',
            })
            counter_writes += 1
        results['id_counters'] = {
            'table_name': get_table_name('id_counters'),
            'written': counter_writes,
            'skipped_existing': 0,
            'total_input': counter_writes,
        }

    return results


def json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    if isinstance(value, dict):
        return {key: json_safe(inner) for key, inner in value.items()}
    if isinstance(value, list):
        return [json_safe(item) for item in value]
    return value


def print_summary(transformed: dict[str, list[dict[str, Any]]], counters: dict[str, int], args: argparse.Namespace, region: str) -> None:
    print('RSA CMS / Mini-CRM Launch Data Import Loader')
    print(f'Region: {region}')
    print(f'Mode: {"EXECUTE" if args.execute else "DRY RUN"}')
    print(f'Source: {args.source}')
    print(f'Existing records: {"OVERWRITE" if args.overwrite else "SKIP"}')
    print(f'Sync id counters: {"YES" if not args.no_sync_counters else "NO"}')
    print('')

    summary = {}
    for table, items in transformed.items():
        pk = primary_key_for_table(table)
        summary[table] = {
            'table_name': get_table_name(table),
            'primary_key': pk,
            'item_count': len(items),
            'sample_keys': [item.get(pk) for item in items[:5]],
        }

    print(json.dumps(json_safe(summary), indent=2))
    print('')
    if not args.execute:
        print('Dry run only. No DynamoDB records were written.')
        print('Run again with --execute when you are ready to import launch data.')


def main() -> int:
    parser = argparse.ArgumentParser(description='Import RSA CMS launch CSV/Excel data into DynamoDB safely.')
    parser.add_argument('--table', default='all', help='Table to import, or all.')
    parser.add_argument('--all', action='store_true', help='Import all supported launch data tables.')
    parser.add_argument('--source', choices=['csv', 'excel'], default='csv', help='Input source type. CSV is default.')
    parser.add_argument('--csv-dir', default=str(DEFAULT_CSV_DIR), help='Folder containing CSV files.')
    parser.add_argument('--excel-file', default=str(DEFAULT_EXCEL_FILE), help='Excel workbook path for --source excel.')
    parser.add_argument('--region', default=get_aws_region(), help='AWS region.')
    parser.add_argument('--execute', action='store_true', help='Actually write records to DynamoDB.')
    parser.add_argument('--overwrite', action='store_true', help='Overwrite existing records instead of skipping.')
    parser.add_argument('--no-sync-counters', action='store_true', help='Do not update rsa_id_counters.')
    args = parser.parse_args()

    selected = 'all' if args.all else args.table
    manifest = load_manifest()
    selected_tables = select_tables(manifest, selected)
    rows_by_table = load_input_rows(args, manifest, selected_tables)

    errors = validate_rows(manifest, rows_by_table)
    if errors:
        print('Validation failed:')
        for error in errors:
            print(f'  ERROR: {error}')
        return 1

    transformed, counters = transform_rows(rows_by_table, region=args.region, execute=args.execute)
    print_summary(transformed, counters, args=args, region=args.region)

    if not args.execute:
        return 0

    results = put_items(
        transformed,
        counters=counters,
        region=args.region,
        overwrite=args.overwrite,
        sync_counters=not args.no_sync_counters,
    )
    print('Import results:')
    print(json.dumps(json_safe(results), indent=2))
    print('')
    print('Launch data import completed.')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
